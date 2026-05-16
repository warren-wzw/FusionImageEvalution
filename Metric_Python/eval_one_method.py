import os
import sys
os.chdir(sys.path[0])

import numpy as np
from PIL import Image
from Metric import *
from natsort import natsorted
from tqdm import tqdm
import warnings
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from concurrent.futures import ProcessPoolExecutor

warnings.filterwarnings("ignore")

DATASET = "LLVIP"
METHOD = ["S2Fusion"]
INFRARED = f"../source_image/{DATASET}/ir"
VISIBLE = f"../source_image/{DATASET}/vi"
ROUND = 6
USE_NR_METRICS = False  # 设为 True 启用 PI/CLIPIQA/MUSIQ（较慢）
NUM_WORKERS = 2  # 并行进程数，设为 1 关闭多进程

if USE_NR_METRICS:
    import pyiqa
    import torch
    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    clipiqa_metric = pyiqa.create_metric('clipiqa', device=device)
    musiq_metric = pyiqa.create_metric('musiq', device=device)
    pi_metric = pyiqa.create_metric('pi', device=device)


def write_all_metrics(excel_name='metric.xlsx', sheet_name='Metrics', metrics_dict=None):
    if metrics_dict is None:
        print("❌ metrics_dict 参数不能为空！")
        return
    try:
        workbook = load_workbook(excel_name)
        if sheet_name in workbook.sheetnames:
            std = workbook[sheet_name]
            workbook.remove(std)
        worksheet = workbook.create_sheet(title=sheet_name)
    except FileNotFoundError:
        workbook = Workbook()
        default_sheet = workbook.active
        workbook.remove(default_sheet)
        worksheet = workbook.create_sheet(title=sheet_name)

    for col_idx, (col_name, col_data) in enumerate(metrics_dict.items()):
        col_letter = get_column_letter(col_idx + 1)
        worksheet[f"{col_letter}1"] = col_name
        for row_idx, value in enumerate(col_data, start=2):
            worksheet[f"{col_letter}{row_idx}"] = value

    workbook.save(excel_name)
    print(f"✅ 成功写入 {excel_name} 的 sheet '{sheet_name}'")


def evaluation_one(ir_name, vi_name, f_name):
    f_img = Image.open(f_name).convert('L')
    ir_img = Image.open(ir_name).convert('L')
    vi_img = Image.open(vi_name).convert('L')

    vi_size = vi_img.size
    if f_img.size != vi_size:
        f_img = f_img.resize(vi_size, Image.BILINEAR)

    f_img_int = np.array(f_img).astype(np.int32)
    f_img_double = np.array(f_img).astype(np.float32)
    ir_img_int = np.array(ir_img).astype(np.int32)
    ir_img_double = np.array(ir_img).astype(np.float32)
    vi_img_int = np.array(vi_img).astype(np.int32)
    vi_img_double = np.array(vi_img).astype(np.float32)

    EN = EN_function(f_img_int)
    MI = MI_function(ir_img_int, vi_img_int, f_img_int, gray_level=256)
    SF = SF_function(f_img_double)
    SD = SD_function(f_img_double)
    AG = AG_function(f_img_double)
    PSNR = PSNR_function(ir_img_double, vi_img_double, f_img_double)
    VIF = VIF_function(ir_img_double, vi_img_double, f_img_double)
    CC = CC_function(ir_img_double, vi_img_double, f_img_double)
    SCD = SCD_function(ir_img_double, vi_img_double, f_img_double)
    Qabf = Qabf_function(ir_img_double, vi_img_double, f_img_double)
    SSIM = SSIM_function(ir_img_double, vi_img_double, f_img_double)
    MS_SSIM = MS_SSIM_function(ir_img_double, vi_img_double, f_img_double)

    # --------------------------------
    # 无参考指标：只对融合图计算
    # --------------------------------
    if USE_NR_METRICS:
        PI = pi_metric(f_name).item()
        CLIPIQA = clipiqa_metric(f_name).item()
        MUSIQ = musiq_metric(f_name).item()
    else:
        PI, CLIPIQA, MUSIQ = None, None, None

    return EN, MI, SF, AG, SD, CC, SCD, VIF, PSNR, Qabf, SSIM, MS_SSIM, PI, CLIPIQA, MUSIQ


def _eval_worker(args):
    os.chdir(sys.path[0])
    ir_name, vi_name, f_name = args
    return evaluation_one(ir_name, vi_name, f_name)


if __name__ == '__main__':
    ir_dir = INFRARED
    vi_dir = VISIBLE

    for method in METHOD:
        Method = method
        f_dir = os.path.join(f'../Results/{DATASET}/', Method)
        save_dir = f'../Excel/{DATASET}'
        os.makedirs(save_dir, exist_ok=True)

        EN_list, MI_list, SF_list, AG_list, SD_list = [], [], [], [], []
        CC_list, SCD_list, VIF_list, PSNR_list = [], [], [], []
        Qabf_list, SSIM_list, MS_SSIM_list = [], [], []
        PI_list, CLIPIQA_list, MUSIQ_list = [], [], []
        filename_list = []

        names = [
            "EN", "MI", "SF", "AG", "SD", "CC", "SCD", "VIF",
            "PSNR", "Qabf", "SSIM", "MS_SSIM"
        ]
        if USE_NR_METRICS:
            names += ["PI", "CLIPIQA", "MUSIQ"]

        name_lists = [
            EN_list, MI_list, SF_list, AG_list, SD_list, CC_list, SCD_list,
            VIF_list, PSNR_list, Qabf_list, SSIM_list, MS_SSIM_list,
        ]
        if USE_NR_METRICS:
            name_lists += [PI_list, CLIPIQA_list, MUSIQ_list]

        metric_save_name = os.path.join(save_dir, f'metric_{Method}.xlsx')
        filelist = natsorted(os.listdir(ir_dir))

        # 过滤掉缺失的融合结果
        valid_items = []
        for item in filelist:
            f_name = os.path.join(f_dir, item)
            if os.path.exists(f_name):
                valid_items.append(item)
            else:
                print(f"⚠️ 缺少融合结果图: {f_name}")

        tasks = [
            (os.path.join(ir_dir, item), os.path.join(vi_dir, item), os.path.join(f_dir, item))
            for item in valid_items
        ]

        use_parallel = not USE_NR_METRICS and NUM_WORKERS > 1
        if use_parallel:
            with ProcessPoolExecutor(max_workers=NUM_WORKERS) as executor:
                results = list(tqdm(
                    executor.map(_eval_worker, tasks),
                    total=len(tasks), desc=Method
                ))
        else:
            results = [evaluation_one(*t) for t in tqdm(tasks, desc=Method)]

        for item, (EN, MI, SF, AG, SD, CC, SCD, VIF, PSNR, Qabf, SSIM, MS_SSIM, PI, CLIPIQA, MUSIQ) in zip(valid_items, results):
            EN_list.append(EN)
            MI_list.append(MI)
            SF_list.append(SF)
            AG_list.append(AG)
            SD_list.append(SD)
            CC_list.append(CC)
            SCD_list.append(SCD)
            VIF_list.append(VIF)
            PSNR_list.append(PSNR)
            Qabf_list.append(Qabf)
            SSIM_list.append(SSIM)
            MS_SSIM_list.append(MS_SSIM)
            PI_list.append(PI)
            CLIPIQA_list.append(CLIPIQA)
            MUSIQ_list.append(MUSIQ)
            filename_list.append(item)

        mean_values = [round(np.mean(lst), ROUND) for lst in name_lists]
        for name_list, mean_val in zip(name_lists, mean_values):
            name_list.insert(0, mean_val)
        filename_list.insert(0, "mean")

        metrics_dict = {"Filename": filename_list}
        for name, name_list in zip(names, name_lists):
            metrics_dict[name] = name_list

        write_all_metrics(
            excel_name=metric_save_name,
            sheet_name=f"{Method}",
            metrics_dict=metrics_dict
        )
# import os
# import sys
# os.chdir(sys.path[0])
# import numpy as np
# from PIL import Image
# from Metric import *
# from natsort import natsorted
# from tqdm import tqdm
# import warnings
# from openpyxl import Workbook, load_workbook
# from openpyxl.utils import get_column_letter
# import pyiqa

# warnings.filterwarnings("ignore")
# DATASET="MSRS"
# METHOD = ["EPOFusion"]
# #"FusionGAN", "GANMcC", "U2Fusion", "SDNet", "TarDal", "SeAFusion", "SegMif", "DDFM", "Dif_Fusion", "Dif_if", "EMMA", "MRFS", "DCINN", "LFDT-Fusion", "A2RNet", "Ours"
# INFRARED=f"../source_image/{DATASET}/ir"
# VISIBLE=f"../source_image/{DATASET}/vi"
# ROUND=6

# def write_all_metrics(excel_name='metric.xlsx', sheet_name='Metrics', metrics_dict=None):
#     if metrics_dict is None:
#         print("❌ metrics_dict 参数不能为空！")
#         return
#     try:
#         workbook = load_workbook(excel_name)
#         if sheet_name in workbook.sheetnames:
#             std = workbook[sheet_name]
#             workbook.remove(std)
#         worksheet = workbook.create_sheet(title=sheet_name)
#     except FileNotFoundError:
#         workbook = Workbook()
#         default_sheet = workbook.active
#         workbook.remove(default_sheet)
#         worksheet = workbook.create_sheet(title=sheet_name)

#     # 写入数据
#     for col_idx, (col_name, col_data) in enumerate(metrics_dict.items()):
#         col_letter = get_column_letter(col_idx + 1)
#         worksheet[f"{col_letter}1"] = col_name  # 第一行写列名
#         for row_idx, value in enumerate(col_data, start=2):
#             worksheet[f"{col_letter}{row_idx}"] = value  # 从第二行开始写值

#     workbook.save(excel_name)
#     print(f"✅ 成功写入 {excel_name} 的 sheet '{sheet_name}'")
    
# def evaluation_one(ir_name, vi_name, f_name):
#     f_img = Image.open(f_name).convert('L')
#     ir_img = Image.open(ir_name).convert('L')
#     vi_img = Image.open(vi_name).convert('L')
#     vi_size = vi_img.size  # (width, height)
#     if f_img.size != vi_size:
#         f_img = f_img.resize(vi_size, Image.BILINEAR)
#     """"""
#     f_img_int = np.array(f_img).astype(np.int32)
#     f_img_double = np.array(f_img).astype(np.float32)
#     ir_img_int = np.array(ir_img).astype(np.int32)
#     ir_img_double = np.array(ir_img).astype(np.float32)
#     vi_img_int = np.array(vi_img).astype(np.int32)
#     vi_img_double = np.array(vi_img).astype(np.float32)

#     EN = EN_function(f_img_int)
#     MI = MI_function(ir_img_int, vi_img_int, f_img_int, gray_level=256)
#     SF = SF_function(f_img_double)
#     SD = SD_function(f_img_double)
#     AG = AG_function(f_img_double)
#     PSNR = PSNR_function(ir_img_double, vi_img_double, f_img_double)
#     VIF = VIF_function(ir_img_double, vi_img_double, f_img_double)
#     CC = CC_function(ir_img_double, vi_img_double, f_img_double)
#     SCD = SCD_function(ir_img_double, vi_img_double, f_img_double)
#     Qabf = Qabf_function(ir_img_double, vi_img_double, f_img_double)
#     SSIM = SSIM_function(ir_img_double, vi_img_double, f_img_double)
#     MS_SSIM = MS_SSIM_function(ir_img_double, vi_img_double, f_img_double)
#     return EN, MI, SF, AG, SD, CC, SCD, VIF, PSNR, Qabf, SSIM, MS_SSIM


# if __name__ == '__main__':
    
#     ir_dir = INFRARED
#     vi_dir = VISIBLE
#     for method in METHOD:
#         Method = method
#         f_dir = os.path.join(f'../Results/{DATASET}/', Method)
#         save_dir = f'../Excel/{DATASET}'
#         os.makedirs(save_dir, exist_ok=True)
#         EN_list, MI_list, SF_list, AG_list, SD_list,CC_list, SCD_list, VIF_list,PSNR_list,Qabf_list,\
#         SSIM_list, MS_SSIM_list, filename_list= [],[],[],[],[],[],[],[],[],[],[],[],[]
#         names=["EN","MI","SF","AG","SD","CC","SCD","VIF","PSNR","Qabf","SSIM","MS_SSIM"]
#         name_lists = [EN_list, MI_list, SF_list, AG_list, SD_list, CC_list, SCD_list, VIF_list,     
#                         PSNR_list, Qabf_list, SSIM_list, MS_SSIM_list]
#         metric_save_name = os.path.join(save_dir, 'metric_{}.xlsx'.format( Method))
#         filelist = natsorted(os.listdir(ir_dir))
#         eval_bar = tqdm(filelist)
#         for _, item in enumerate(eval_bar):
#             ir_name = os.path.join(ir_dir, item)
#             vi_name = os.path.join(vi_dir, item)
#             f_name = os.path.join(f_dir, item)
#             EN, MI, SF, AG, SD, CC, SCD, VIF, PSNR, Qabf, SSIM, MS_SSIM = evaluation_one(ir_name, vi_name,f_name)
#             EN_list.append(EN)
#             MI_list.append(MI)
#             SF_list.append(SF)
#             AG_list.append(AG)
#             SD_list.append(SD)
#             CC_list.append(CC)
#             SCD_list.append(SCD)
#             VIF_list.append(VIF)
#             PSNR_list.append(PSNR)
#             Qabf_list.append(Qabf)
#             SSIM_list.append(SSIM)
#             MS_SSIM_list.append(MS_SSIM)
#             filename_list.append(item)
#             eval_bar.set_description("{} | {}".format(Method, item))
#         mean_values = [round(np.mean(lst), ROUND) for lst in name_lists]
#         for name_list, mean_val in zip(name_lists, mean_values):
#             name_list.insert(0, mean_val)
#         filename_list.insert(0, "mean")
#         metrics_dict = {"Filename": filename_list}
#         for name, name_list in zip(names, name_lists):
#             metrics_dict[name] = name_list
#         write_all_metrics(
#             excel_name=metric_save_name,
#             sheet_name=f"{Method}",
#             metrics_dict=metrics_dict
#         )
